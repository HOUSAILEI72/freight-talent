"""
import_taxonomy.py — Excel 解析（taxonomy 模型）

模型：
  - 列名同义词表 → 命中固定字段（写入 jobs/candidates 主表字段）
  - 未命中 → 列名原样作为 category；单元格值（支持多值）作为 tag
  - 长文本列（超过 LONG_TEXT_LIMIT 或含换行）默认不进 tag
  - "省市区" 字段特殊：用 - 拆 province / city_name / district

输出：
  TaxonomyPreview {
      rows:           list[ParsedRow]
      categories:     list[CategoryStat]    # 全部 category 维度统计
      summary:        dict
      issues:         list[dict]            # 文件级问题
  }

  ParsedRow {
      row_index:      int      # 1-based Excel 行号（含表头则数据从 2 起）
      data_fields:    dict     # 命中固定字段后规范化的值
      tags:           list[(category, name)]
      issues:         list[dict]
      raw_data:       dict     # 原始单元格快照
      fingerprint:    str      # 用于重复检测
  }

调用方负责：
  1. 把 ParsedRow 写入 import_batch_rows（已有逻辑）
  2. 把 tags 字段铺平写入 import_batch_tags
  3. confirm 时根据 ParsedRow.data_fields 写主表，再用 batch_tags 建 junction
"""
from __future__ import annotations

import hashlib
import io
import re
from dataclasses import dataclass, field
from typing import Any

from openpyxl import load_workbook


# ── 常量 ──────────────────────────────────────────────────────────────────────

LONG_TEXT_LIMIT = 60        # 单元格超过此长度不当 tag
MAX_ROWS = 5000
MAX_FILE_BYTES = 20 * 1024 * 1024   # 20 MB
TAG_SPLIT_RE = re.compile(r"[,，、;；/]")     # 单元格内多值分隔符


# 固定字段定义 — 列头别名 → 标准 field_key
JOB_FIELD_ALIASES: dict[str, str] = {
    "岗位名称": "title", "职位名称": "title", "职位": "title", "岗位": "title", "title": "title",
    "省市区": "location", "工作地点": "location", "地点": "location", "location": "location",
    "工作城市": "location",
    "经验要求": "experience_required", "工作经验": "experience_required",
    "工作年限": "experience_required", "experience_required": "experience_required",
    "最低学历要求": "degree_required",
    "学历": "degree_required", "学历要求": "degree_required", "degree_required": "degree_required",
    "薪资": "salary_label", "薪资范围": "salary_label", "薪酬": "salary_label",
    "工资": "salary_label", "salary": "salary_label", "salary_label": "salary_label",
    "职位详情": "description", "岗位描述": "description", "职位描述": "description",
    "岗位职责": "description", "工作内容": "description", "description": "description",
    "岗位要求": "requirements", "任职要求": "requirements", "职位要求": "requirements",
    "requirements": "requirements",
    "企业邮箱": "company_email", "company_email": "company_email",
    # ── Phase C / PostJob 新增字段 ──────────────────────────────────────────
    "岗位板块": "function_code", "板块": "function_code", "function_code": "function_code",
    "知识": "knowledge_requirements", "知识要求": "knowledge_requirements",
    "硬技能": "hard_skill_requirements", "硬技能要求": "hard_skill_requirements",
    "软技能": "soft_skill_requirements", "软技能要求": "soft_skill_requirements",
    "是否带团队": "is_management_role", "带团队": "is_management_role",
    "预计团队人数": "management_headcount", "团队人数": "management_headcount",
    "应聘类型": "employment_type", "雇佣类型": "employment_type",
    "最低月薪": "salary_min", "月薪下限": "salary_min",
    "最高月薪": "salary_max", "月薪上限": "salary_max",
    "薪资月数": "salary_months", "月薪月数": "salary_months",
    "提成奖金周期": "commission_bonus_period", "提成周期": "commission_bonus_period",
    "提成预估平均额": "commission_bonus_amount", "提成平均额": "commission_bonus_amount",
    "是否有年终奖": "has_year_end_bonus", "年终奖": "has_year_end_bonus",
    "年终奖月数": "year_end_bonus_months",
    "详细地址": "address_detail",
}

CANDIDATE_FIELD_ALIASES: dict[str, str] = {
    "姓名": "full_name", "全名": "full_name", "候选人姓名": "full_name", "full_name": "full_name",
    "年龄": "age", "age": "age",
    "工作年限": "experience_years", "工作经验": "experience_years",
    "经验年限": "experience_years", "experience_years": "experience_years",
    "学历": "education", "最高学历": "education", "education": "education",
    "电话号码": "phone", "电话": "phone", "手机": "phone", "联系电话": "phone", "phone": "phone",
    "求职状态": "availability_status", "在职状态": "availability_status",
    "availability_status": "availability_status",
    "工作经历": "work_experiences", "work_experiences": "work_experiences",
    "教育经历": "education_experiences", "education_experiences": "education_experiences",
    "资格证书": "certificates", "certificates": "certificates",
    "邮箱": "email", "联系邮箱": "email", "email": "email",
    "性别": "gender", "gender": "gender",
    # ── Phase C / CandidateDetailPanel 新增字段 ──────────────────────────────
    "岗位板块": "function_code", "板块": "function_code",
    "是否管理岗": "is_management_role", "带团队": "is_management_role",
    "管理人数": "management_headcount",
    "知识标签": "knowledge_tags", "知识": "knowledge_tags",
    "硬技能标签": "hard_skill_tags", "硬技能": "hard_skill_tags",
    "软技能标签": "soft_skill_tags", "软技能": "soft_skill_tags",
    "期望城市": "expected_city", "意向城市": "expected_city",
    "期望薪资": "expected_salary_label", "期望薪酬": "expected_salary_label",
    "个人简介": "summary", "简介": "summary",
    "当前职位": "current_title", "现任职位": "current_title",
    "当前公司": "current_company", "现任公司": "current_company",
    "当前职责": "current_responsibilities",
    "当前月薪下限": "current_salary_min",
    "当前月薪上限": "current_salary_max",
    "当前薪资月数": "current_salary_months",
    "当前奖金比例": "current_average_bonus_percent",
    "当前是否有年终奖": "current_has_year_end_bonus",
    "当前年终奖月数": "current_year_end_bonus_months",
}

# 这些固定字段同时也生成 tag（默认开启）
DOUBLE_AS_TAG_JOB = {"experience_required", "degree_required"}
DOUBLE_AS_TAG_CANDIDATE = {"education", "availability_status"}

# 提成奖金周期：中文 → DB enum 值
COMMISSION_PERIOD_MAP: dict[str, str] = {
    "不适用": "not_applicable",
    "无": "not_applicable",
    "月度": "monthly",
    "每月": "monthly",
    "季度": "quarterly",
    "每季": "quarterly",
    "半年度": "semi_annual",
    "半年": "semi_annual",
    # 英文直通（兼容旧模板）
    "not_applicable": "not_applicable",
    "monthly": "monthly",
    "quarterly": "quarterly",
    "semi_annual": "semi_annual",
}

# 岗位板块：中/英文 → function_code key（对应 FunctionRail.DEFAULT_FUNCTIONS）
FUNCTION_CODE_MAP: dict[str, str] = {
    # 海运板块
    "海运": "Sea", "Sea": "Sea", "海运板块": "Sea",
    # 空运板块
    "空运": "Air", "Air": "Air", "空运板块": "Air",
    # 跨境电商物流
    "跨境电商": "CrossBorder", "CrossBorder": "CrossBorder", "ECOMS": "CrossBorder",
    "综合物流": "CrossBorder", "跨境电商物流": "CrossBorder",
    # 铁路 / 中欧班列
    "铁路": "Railway", "Railway": "Railway", "中欧班列": "Railway",
    "多式联运": "Railway", "MultiModal": "Railway",
    # 陆路运输
    "陆运": "Road", "Road": "Road", "陆路运输": "Road",
    "Land": "Road",
    # 合同物流 / 3PL
    "合同物流": "ContractLogistics", "ContractLogistics": "ContractLogistics",
    "Contract Logistics": "ContractLogistics", "CL": "ContractLogistics",
    "3PL": "ContractLogistics", "供应链": "ContractLogistics", "Supply": "ContractLogistics",
    # 仓储 / 海外仓
    "仓储": "Warehousing", "Warehousing": "Warehousing", "海外仓": "Warehousing",
    # 关务 / 合规
    "报关": "Customs", "Customs": "Customs", "关务": "Customs",
    "合规": "Customs", "Custom": "Customs", "通关": "Customs",
}


# ── 数据结构 ──────────────────────────────────────────────────────────────────

@dataclass
class ParsedRow:
    row_index: int
    data_fields: dict[str, Any] = field(default_factory=dict)
    tags: list[tuple[str, str]] = field(default_factory=list)
    issues: list[dict] = field(default_factory=list)
    raw_data: dict = field(default_factory=dict)
    fingerprint: str = ""


@dataclass
class CategoryStat:
    category: str
    is_new: bool
    tag_count: int
    tags: list[dict]  # [{name, count, is_new}]


@dataclass
class TaxonomyPreview:
    headers: list[str]
    fixed_field_map: dict[str, str]   # column_header → field_key（已命中固定字段的列）
    free_columns: list[str]        # 未命中固定字段的列名（即 category）
    rows: list[ParsedRow] = field(default_factory=list)
    categories: list[CategoryStat] = field(default_factory=list)
    summary: dict = field(default_factory=dict)
    issues: list[dict] = field(default_factory=list)


# ── 工具 ──────────────────────────────────────────────────────────────────────

def _norm(s: Any) -> str:
    if s is None:
        return ""
    return str(s).strip()


def _split_multi(val: str) -> list[str]:
    parts = [p.strip() for p in TAG_SPLIT_RE.split(val) if p.strip()]
    return parts


def _is_long_text(val: str) -> bool:
    return len(val) > LONG_TEXT_LIMIT or "\n" in val


def _safe_int(val: Any) -> int | None:
    if val is None or val == "":
        return None
    try:
        return int(float(val))
    except (TypeError, ValueError):
        return None


def _safe_float(val: Any) -> float | None:
    if val is None or val == "":
        return None
    try:
        return float(val)
    except (TypeError, ValueError):
        return None


def _parse_location(val: str) -> tuple[str | None, str | None, str | None]:
    """'上海市-上海市-浦东新区' → (province, city, district)；缺位补 None。"""
    if not val:
        return None, None, None
    parts = [p.strip() for p in re.split(r"[-—–]", val) if p.strip()]
    while len(parts) < 3:
        parts.append(None)
    return parts[0], parts[1], parts[2]


def _parse_work_exp(cell: str) -> list[dict]:
    """每行一条；字段用 | 分隔：年限 | 职位 | 公司"""
    out = []
    for line in (cell or "").split("\n"):
        line = line.strip()
        if not line:
            continue
        parts = [p.strip() for p in line.split("|")]
        out.append({
            "period": parts[0] if len(parts) > 0 else "",
            "title": parts[1] if len(parts) > 1 else "",
            "company": parts[2] if len(parts) > 2 else "",
        })
    return out


def _parse_edu_exp(cell: str) -> list[dict]:
    """每行一条；字段用 | 分隔：年限 | 院校 | 专业 | 学历"""
    out = []
    for line in (cell or "").split("\n"):
        line = line.strip()
        if not line:
            continue
        parts = [p.strip() for p in line.split("|")]
        out.append({
            "period": parts[0] if len(parts) > 0 else "",
            "school": parts[1] if len(parts) > 1 else "",
            "major": parts[2] if len(parts) > 2 else "",
            "degree": parts[3] if len(parts) > 3 else "",
        })
    return out


def _parse_certificates(cell: str) -> list[str]:
    return [p.strip() for p in re.split(r"[,，、;；\n]", cell or "") if p.strip()]


def _fingerprint(data_fields: dict, raw_data: dict) -> str:
    """用关键字段生成行指纹用于重复检测。"""
    parts = []
    for k in sorted(raw_data.keys()):
        parts.append(f"{k}={_norm(raw_data[k])}")
    return hashlib.sha256("\n".join(parts).encode("utf-8")).hexdigest()


# ── 主入口 ────────────────────────────────────────────────────────────────────

def parse_excel(
    file_bytes: bytes,
    import_type: str,
    known_categories: set[str] | None = None,
    known_tags: set[tuple[str, str]] | None = None,
) -> TaxonomyPreview:
    """
    解析 Excel，返回 TaxonomyPreview。

    Args:
      import_type:        "job" | "resume"
      known_categories:   已存在的 category 名称集合（任意 source 的 active tag）
      known_tags:         已存在的 (category, name) 集合
    """
    known_categories = known_categories or set()
    known_tags = known_tags or set()

    aliases = JOB_FIELD_ALIASES if import_type == "job" else CANDIDATE_FIELD_ALIASES
    double_as_tag = DOUBLE_AS_TAG_JOB if import_type == "job" else DOUBLE_AS_TAG_CANDIDATE

    preview = TaxonomyPreview(headers=[], fixed_field_map={}, free_columns=[])

    # ── 1. 文件解析 ──────────────────────────────────────────────────────────
    try:
        wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        all_rows = list(rows_iter)
    except Exception as e:
        preview.issues.append({"issue_type": "file_parse_error", "suggestion": f"无法解析 Excel：{e}"})
        return preview

    if not all_rows:
        preview.issues.append({"issue_type": "empty_file", "suggestion": "Excel 文件为空"})
        return preview

    headers_raw = [_norm(c) for c in all_rows[0]]
    preview.headers = headers_raw

    # 列头去重检测
    seen = {}
    dup_indices = set()
    for i, h in enumerate(headers_raw):
        if not h:
            continue
        if h in seen:
            dup_indices.add(i)
            dup_indices.add(seen[h])
        else:
            seen[h] = i
    if dup_indices:
        preview.issues.append({
            "issue_type": "duplicate_header",
            "columns": [headers_raw[i] for i in sorted(dup_indices)],
            "suggestion": "存在重复列头，请合并后重新上传",
        })
        return preview

    # ── 2. 列归类：固定字段 vs 自由 category ──────────────────────────────────
    for h in headers_raw:
        if not h:
            continue
        fk = aliases.get(h)
        if fk:
            preview.fixed_field_map[h] = fk
        else:
            preview.free_columns.append(h)

    # ── 3. 行解析 ────────────────────────────────────────────────────────────
    data_rows = all_rows[1:]
    if len(data_rows) > MAX_ROWS:
        preview.issues.append({
            "issue_type": "row_limit",
            "suggestion": f"超出 {MAX_ROWS} 行上限，仅解析前 {MAX_ROWS} 行",
        })
        data_rows = data_rows[:MAX_ROWS]

    cat_tag_counts: dict[str, dict[str, int]] = {}    # category → tag → count

    for ridx, row_vals in enumerate(data_rows, start=2):  # Excel 第 2 行起
        raw = {h: _norm(v) for h, v in zip(headers_raw, row_vals) if h}
        parsed = ParsedRow(row_index=ridx, raw_data=raw)

        # 3.1 固定字段
        for h, fk in preview.fixed_field_map.items():
            v = raw.get(h, "")
            if not v:
                continue
            if fk == "location":
                p, c, d = _parse_location(v)
                if p:
                    parsed.data_fields["province"] = p
                if c:
                    parsed.data_fields["city_name"] = c
                if d:
                    parsed.data_fields["district"] = d
                if c:
                    parsed.data_fields["city"] = c     # 兼容旧字段
                # 三级分别变 tag（同分类 OR）
                for level, val in [("省份", p), ("城市", c), ("区县", d)]:
                    if val:
                        cat_tag_counts.setdefault(level, {}).setdefault(val, 0)
                        cat_tag_counts[level][val] += 1
                        parsed.tags.append((level, val))
                continue
            if fk == "work_experiences":
                parsed.data_fields[fk] = _parse_work_exp(v)
                continue
            if fk == "education_experiences":
                parsed.data_fields[fk] = _parse_edu_exp(v)
                continue
            if fk == "certificates":
                parsed.data_fields[fk] = _parse_certificates(v)
                continue
            if fk in ("age", "experience_years", "headcount",
                      "management_headcount",
                      "salary_min", "salary_max", "salary_months",
                      "commission_bonus_amount", "year_end_bonus_months",
                      "current_salary_min", "current_salary_max", "current_salary_months",
                      "current_year_end_bonus_months", "current_average_bonus_percent"):
                iv = _safe_int(v) if fk not in ("commission_bonus_amount",
                                                "year_end_bonus_months",
                                                "current_year_end_bonus_months",
                                                "current_average_bonus_percent") else _safe_float(v)
                if iv is not None:
                    parsed.data_fields[fk] = iv
                else:
                    parsed.issues.append({"field": fk, "issue_type": "type_mismatch",
                                          "original_value": v, "suggestion": "应为数字"})
                continue
            if fk in ("is_management_role", "has_year_end_bonus", "current_has_year_end_bonus"):
                parsed.data_fields[fk] = v.lower() in ("是", "yes", "true", "1")
                continue
            if fk in ("knowledge_requirements", "hard_skill_requirements", "soft_skill_requirements",
                      "knowledge_tags", "hard_skill_tags", "soft_skill_tags"):
                parsed.data_fields[fk] = _split_multi(v) if v else []
                continue
            if fk == "commission_bonus_period":
                parsed.data_fields[fk] = COMMISSION_PERIOD_MAP.get(v, v)
                continue
            if fk == "function_code":
                mapped = FUNCTION_CODE_MAP.get(v, v)
                parsed.data_fields[fk] = mapped
                parsed.data_fields["function_name"] = mapped
                continue
            parsed.data_fields[fk] = v
            # 双轨：固定字段也作为 tag
            if fk in double_as_tag:
                cat_label = h    # 用列头中文作为 category
                cat_tag_counts.setdefault(cat_label, {}).setdefault(v, 0)
                cat_tag_counts[cat_label][v] += 1
                parsed.tags.append((cat_label, v))

        # 3.2 自由列 → category + tag
        for col in preview.free_columns:
            v = raw.get(col, "")
            if not v or _is_long_text(v):
                continue
            for piece in _split_multi(v) or [v]:
                cat_tag_counts.setdefault(col, {}).setdefault(piece, 0)
                cat_tag_counts[col][piece] += 1
                parsed.tags.append((col, piece))

        # 3.3 必填校验
        if import_type == "job":
            if not parsed.data_fields.get("title"):
                parsed.issues.append({"field": "title", "issue_type": "missing_required",
                                      "suggestion": "缺少岗位名称"})
            if not parsed.data_fields.get("city_name") and not parsed.data_fields.get("city"):
                parsed.issues.append({"field": "city", "issue_type": "missing_required",
                                      "suggestion": "缺少省市区"})
        else:
            if not parsed.data_fields.get("full_name"):
                parsed.issues.append({"field": "full_name", "issue_type": "missing_required",
                                      "suggestion": "缺少姓名"})
            if not parsed.data_fields.get("phone") and not parsed.data_fields.get("email"):
                parsed.issues.append({"field": "contact", "issue_type": "missing_required",
                                      "suggestion": "缺少电话或邮箱（至少一项）"})

        parsed.fingerprint = _fingerprint(parsed.data_fields, raw)
        preview.rows.append(parsed)

    # ── 4. 重复行检测 ────────────────────────────────────────────────────────
    fp_seen: dict[str, int] = {}
    for p in preview.rows:
        if p.fingerprint in fp_seen:
            p.issues.append({"issue_type": "duplicate_row",
                             "suggestion": f"与第 {fp_seen[p.fingerprint]} 行内容重复"})
        else:
            fp_seen[p.fingerprint] = p.row_index

    # ── 5. 聚合 category 统计 ────────────────────────────────────────────────
    for cat, tag_counts in sorted(cat_tag_counts.items()):
        is_new_cat = cat not in known_categories
        tags_stat = []
        for name, cnt in sorted(tag_counts.items(), key=lambda x: -x[1]):
            tags_stat.append({
                "name": name,
                "count": cnt,
                "is_new": (cat, name) not in known_tags,
            })
        preview.categories.append(CategoryStat(
            category=cat,
            is_new=is_new_cat,
            tag_count=sum(tag_counts.values()),
            tags=tags_stat,
        ))

    # ── 6. summary ───────────────────────────────────────────────────────────
    total_rows = len(preview.rows)
    error_rows = sum(1 for p in preview.rows if any(
        i.get("issue_type") in ("missing_required", "type_mismatch") for i in p.issues))
    dup_rows = sum(1 for p in preview.rows if any(
        i.get("issue_type") == "duplicate_row" for i in p.issues))
    ok_rows = total_rows - error_rows - dup_rows

    preview.summary = {
        "total_rows": total_rows,
        "ok_rows": ok_rows,
        "error_rows": error_rows,
        "warning_rows": 0,
        "dup_rows": dup_rows,
        "fixed_columns_hit": len(preview.fixed_field_map),
        "free_categories": len(preview.free_columns),
        "new_categories": sum(1 for c in preview.categories if c.is_new),
        "new_tags": sum(
            sum(1 for t in c.tags if t["is_new"]) for c in preview.categories
        ),
    }

    return preview
