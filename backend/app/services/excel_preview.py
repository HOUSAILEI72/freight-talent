"""
excel_preview.py — Excel 导入预检服务

职责：
  1. 从上传的 .xlsx/.xls 文件中读取数据
  2. 识别列头、推断每列类型（text/number/date/datetime/boolean）
  3. 检测类型冲突（同列存在不兼容大类）
  4. 检测完全重复行（基于规范化指纹）
  5. 检测纯文本全局重复（长度 >= 8 的 text 单元格跨行跨列）
  6. 识别新字段（对比字段注册表）
  7. 生成带颜色标注的 Excel 文件（含 Issues sheet）
  8. 返回结构化预检结果

颜色语义：
  红色   — 类型冲突
  橙色   — 纯文本全局重复
  黄色   — 整行重复（重复岗位/记录）
  紫色   — 重复列名（表头级错误）
"""
from __future__ import annotations

import hashlib
import io
import re
import unicodedata
from datetime import date, datetime, time
from typing import Any

import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

# ── 颜色常量 ────────────────────────────────────────────────────────────────
_FILL_RED    = PatternFill("solid", fgColor="FFCCCC")   # 类型冲突
_FILL_ORANGE = PatternFill("solid", fgColor="FFD966")   # 纯文本全局重复（橙黄）
_FILL_YELLOW = PatternFill("solid", fgColor="FFFF99")   # 整行重复
_FILL_PURPLE = PatternFill("solid", fgColor="E0AAFF")   # 重复列名（表头级错误）

# ── 允许的字段大类 ──────────────────────────────────────────────────────────
ALLOWED_TYPES = {"text", "number", "date", "datetime", "boolean"}

# ── 最大行数保护（防止超大文件压垮服务）───────────────────────────────────
MAX_ROWS = 5000

# ── 中文列名 → field_key 别名映射 ────────────────────────────────────────────
# 支持常见中文表头写法，预检时先做替换再走后续逻辑
_CN_ALIAS: dict[str, str] = {
    # ── 岗位字段 ──────────────────────────────────────────────────────────────
    "职位名称": "title",
    "岗位名称": "title",
    "职位":     "title",
    "岗位":     "title",
    "城市":     "city",
    "工作城市": "city",
    "所在城市": "city",
    "职位描述": "description",
    "岗位描述": "description",
    "工作描述": "description",
    "职位详情": "description",
    "描述":     "description",
    "业务类型": "business_type",
    "岗位类型": "job_type",
    "职位类型": "job_type",
    "薪资":     "salary_label",
    "薪资范围": "salary_label",
    "薪酬":     "salary_label",
    "工资":     "salary_label",
    "经验要求": "experience_required",
    "工作年限": "experience_required",
    "学历要求": "degree_required",
    "学历":     "degree_required",
    "招聘人数": "headcount",
    "人数":     "headcount",
    "职位要求": "requirements",
    "任职要求": "requirements",
    "航线标签": "route_tags",
    "航线":     "route_tags",
    "技能标签": "skill_tags",
    "技能要求": "skill_tags",
    "技能":     "skill_tags",
    "紧急程度": "urgency_level",
    "状态":     "status",
    "企业邮箱": "company_email",
    # ── 简历字段 ──────────────────────────────────────────────────────────────
    "姓名":     "full_name",
    "全名":     "full_name",
    "候选人姓名": "full_name",
    "当前职位": "current_title",
    "现任职位": "current_title",
    "职位头衔": "current_title",
    "当前城市": "current_city",
    "现居城市": "current_city",
    "常住城市": "current_city",
    "当前公司": "current_company",
    "现任公司": "current_company",
    "公司":     "current_company",
    "期望城市": "expected_city",
    "意向城市": "expected_city",
    "期望薪资": "expected_salary_label",
    "期望薪酬": "expected_salary_label",
    "工作年限": "experience_years",
    "工作经验": "experience_years",
    "经验年限": "experience_years",
    "学历":     "education",
    "最高学历": "education",
    "英语水平": "english_level",
    "英文水平": "english_level",
    "个人简介": "summary",
    "自我介绍": "summary",
    "简介":     "summary",
    "求职状态": "availability_status",
    "在职状态": "availability_status",
    "邮箱":     "email",
    "联系邮箱": "email",
    "电话":     "phone",
    "手机":     "phone",
    "联系电话": "phone",
    "地址":     "address",
    "常住地址": "address",
}



# ===========================================================================
# 类型推断
# ===========================================================================

def _infer_cell_type(value: Any) -> str | None:
    """
    推断单个单元格的值属于哪个大类。
    返回 None 表示空值，跳过类型冲突判断。
    """
    if value is None or (isinstance(value, str) and value.strip() == ""):
        return None
    if isinstance(value, bool):
        return "boolean"
    if isinstance(value, (int, float)):
        return "number"
    if isinstance(value, datetime):
        return "datetime"
    if isinstance(value, date):
        return "date"
    if isinstance(value, time):
        return "text"   # time-only 当 text 处理
    if isinstance(value, str):
        s = value.strip()
        # 尝试 boolean 字面量
        if s.lower() in ("true", "false", "yes", "no", "是", "否", "1", "0"):
            # 不强制判断为 boolean，保留为 text，因为这类字段通常按文本处理
            pass
        # 尝试 datetime
        if _try_parse_datetime(s):
            return "datetime"
        # 尝试 date
        if _try_parse_date(s):
            return "date"
        # 尝试 number
        try:
            float(s.replace(",", "").replace("，", ""))
            return "number"
        except ValueError:
            pass
        return "text"
    return "text"


_DATE_PATTERNS = [
    re.compile(r"^\d{4}[-/年]\d{1,2}[-/月]\d{1,2}日?$"),
    re.compile(r"^\d{1,2}[-/]\d{1,2}[-/]\d{4}$"),
]
_DATETIME_PATTERNS = [
    re.compile(r"^\d{4}[-/年]\d{1,2}[-/月]\d{1,2}[日\s]\s*\d{1,2}:\d{2}"),
]


def _try_parse_date(s: str) -> bool:
    return any(p.match(s) for p in _DATE_PATTERNS)


def _try_parse_datetime(s: str) -> bool:
    return any(p.match(s) for p in _DATETIME_PATTERNS)


# ===========================================================================
# 规范化 & 指纹
# ===========================================================================

def _normalize_text(v: Any) -> str:
    """将单元格值规范化为可比较字符串。"""
    if v is None:
        return ""
    if isinstance(v, bool):
        return "true" if v else "false"
    if isinstance(v, (int, float)):
        # 统一浮点展示（去掉无用的 .0）
        if isinstance(v, float) and v == int(v):
            return str(int(v))
        return str(v)
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(v, date):
        return v.strftime("%Y-%m-%d")
    # str：去首尾空格、统一连续空白、统一换行、unicode NFC
    s = unicodedata.normalize("NFC", str(v))
    s = s.strip()
    s = re.sub(r"\r\n|\r", "\n", s)
    s = re.sub(r"[ \t]+", " ", s)
    return s


def _row_fingerprint(row_values: list[Any], headers: list[str]) -> str:
    """
    计算一行的规范化指纹。
    标签类字段（字段名含 tag/tags）内部排序后参与 hash。
    """
    parts = []
    for h, v in zip(headers, row_values):
        norm = _normalize_text(v)
        # 标签字段：排序后拼接
        if "tag" in h.lower():
            items = [x.strip() for x in re.split(r"[,，、]", norm) if x.strip()]
            norm = ",".join(sorted(items))
        parts.append(f"{h}={norm}")
    raw = "|".join(parts)
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()


# ===========================================================================
# 核心预检逻辑
# ===========================================================================

class PreviewResult:
    """预检结果容器。"""
    def __init__(self):
        self.headers: list[str] = []
        self.inferred_types: dict[str, str] = {}         # header -> type
        self.type_conflicts: dict[str, list[str]] = {}   # header -> [type1, type2, ...]
        self.new_fields: list[dict] = []                  # 未在字段注册表中的列
        self.row_results: list[dict] = []                 # 每行的预检结果
        self.total_rows: int = 0
        self.ok_rows: int = 0
        self.error_rows: int = 0
        self.warning_rows: int = 0
        self.dup_rows: int = 0
        self.errors: list[dict] = []
        self.warnings: list[dict] = []


def run_preview(
    file_bytes: bytes,
    import_type: str,
    known_field_keys: set[str],
) -> PreviewResult:
    """
    执行预检，返回 PreviewResult。

    参数：
      file_bytes      — Excel 文件内容
      import_type     — "job" | "resume"
      known_field_keys — 字段注册表中已有的 field_key 集合（用于识别新字段）
    """
    result = PreviewResult()

    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    except Exception as exc:
        result.errors.append({
            "issue_type": "file_parse_error",
            "suggestion": f"无法解析 Excel 文件：{exc}",
        })
        return result

    ws = wb.active

    # ── 读取表头 ─────────────────────────────────────────────────────────────
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_row:
        result.errors.append({
            "issue_type": "no_header",
            "suggestion": "Excel 文件首行为空，无法识别列头",
        })
        return result

    headers = []
    _seen_hdr: dict[str, int] = {}   # normalized_name -> first col_idx (0-based)
    _dup_col_indices: set[int] = set()  # col_idx 属于重复列头，不进 field_registry

    for cell_val in header_row:
        h = _normalize_text(cell_val)
        if not h:
            h = f"__col_{len(headers)}"
        col_idx = len(headers)

        if h in _seen_hdr:
            # 记录重复：全局 error（非行级），标记该列不进 field_registry
            _dup_col_indices.add(col_idx)
            result.errors.append({
                "issue_type": "duplicate_header",
                "field": h,
                "col_index": col_idx + 1,   # 1-based，方便 UI 定位
                "suggestion": (
                    f"第 {col_idx + 1} 列列头 '{h}' 与第 {_seen_hdr[h] + 1} 列重复，"
                    "重复列数据已保留但不会注册为独立字段，请在原文件中修正列名"
                ),
            })
            # 仍然追加带后缀的名字，保留列数据供 raw_data 审计
            suffix_count = sum(1 for x in headers if x == h or x.startswith(f"{h}__"))
            h = f"{h}__{suffix_count + 1}"
        else:
            _seen_hdr[h] = col_idx

        headers.append(h)

    result.headers = headers

    # ── 中文列名别名替换 ───────────────────────────────────────────────────────
    # 将识别到的中文列头替换为对应的英文 field_key，后续所有逻辑统一走英文 key
    headers = [_CN_ALIAS.get(h, h) for h in headers]
    result.headers = headers

    # ── 读取数据行 ────────────────────────────────────────────────────────────
    data_rows: list[list[Any]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_vals = list(row[: len(headers)])
        # 补齐列数不足的行
        while len(row_vals) < len(headers):
            row_vals.append(None)
        # 跳过完全空行
        if all(v is None or (isinstance(v, str) and v.strip() == "") for v in row_vals):
            continue
        data_rows.append(row_vals)
        if len(data_rows) >= MAX_ROWS:
            result.warnings.append({
                "issue_type": "row_limit_truncated",
                "suggestion": f"文件超过 {MAX_ROWS} 行，已截断处理，剩余行未检查",
            })
            break

    result.total_rows = len(data_rows)

    if result.total_rows == 0:
        result.warnings.append({
            "issue_type": "no_data",
            "suggestion": "Excel 文件没有数据行（仅有表头）",
        })
        return result

    # ── 按列收集所有非空值并推断类型 ─────────────────────────────────────────
    col_types: dict[str, set[str]] = {h: set() for h in headers}
    for row_vals in data_rows:
        for h, v in zip(headers, row_vals):
            t = _infer_cell_type(v)
            if t is not None:
                col_types[h].add(t)

    for h, types in col_types.items():
        if len(types) == 0:
            result.inferred_types[h] = "text"   # 全空列默认 text
        elif len(types) == 1:
            result.inferred_types[h] = next(iter(types))
        else:
            # 类型冲突
            result.inferred_types[h] = "text"   # 降级
            result.type_conflicts[h] = sorted(types)

    # ── 识别新字段 ────────────────────────────────────────────────────────────
    # 排除重复列头（_dup_col_indices）和已在 known_field_keys 中的字段
    _seen_new_fk: set[str] = set()
    for col_idx, h in enumerate(headers):
        if col_idx in _dup_col_indices:
            continue   # 重复列不进 field_registry，避免 synthetic key 污染
        fk = _to_field_key(h)
        if fk not in known_field_keys and fk not in _seen_new_fk:
            _seen_new_fk.add(fk)
            result.new_fields.append({
                "field_key": fk,
                "label": h,
                "inferred_type": result.inferred_types.get(h, "text"),
            })

    # ── 纯文本全局重复检测 ───────────────────────────────────────────────────
    # 收集所有 text 类型列中 len >= 8 的值，记录 (row_idx, col_idx, value)
    text_value_positions: dict[str, list[tuple[int, int]]] = {}  # value -> [(row_idx, col_idx)]
    for row_idx, row_vals in enumerate(data_rows):
        for col_idx, (h, v) in enumerate(zip(headers, row_vals)):
            if result.inferred_types.get(h) == "text" and isinstance(v, str):
                s = v.strip()
                if len(s) >= 8:
                    text_value_positions.setdefault(s, []).append((row_idx, col_idx))

    # 出现超过 1 次的文本值
    global_text_dups: dict[tuple[int, int], str] = {}  # (row_idx, col_idx) -> dup_value
    for val, positions in text_value_positions.items():
        if len(positions) > 1:
            for pos in positions:
                global_text_dups[pos] = val

    # ── 完全重复行检测（基于规范化指纹）────────────────────────────────────
    seen_fingerprints: dict[str, int] = {}   # fingerprint -> first row_idx
    dup_row_indices: set[int] = set()
    row_fingerprints: list[str] = []

    for row_idx, row_vals in enumerate(data_rows):
        fp = _row_fingerprint(row_vals, headers)
        row_fingerprints.append(fp)
        if fp in seen_fingerprints:
            dup_row_indices.add(row_idx)
        else:
            seen_fingerprints[fp] = row_idx

    # ── 组装行级结果 ─────────────────────────────────────────────────────────
    for row_idx, row_vals in enumerate(data_rows):
        issues = []
        row_status = "ok"
        excel_row_num = row_idx + 2   # 1-based，含表头行偏移

        # 1) 类型冲突 — 标注本行中属于冲突列的单元格
        for col_idx, h in enumerate(headers):
            if h in result.type_conflicts:
                cell_type = _infer_cell_type(row_vals[col_idx])
                if cell_type is not None:   # 非空才标注
                    issues.append({
                        "row": excel_row_num,
                        "field": h,
                        "issue_type": "type_conflict",
                        "original_value": str(row_vals[col_idx]),
                        "suggestion": (
                            f"该列存在类型冲突 {result.type_conflicts[h]}，"
                            f"当前单元格推断为 {cell_type}"
                        ),
                    })
                    row_status = _worse_status(row_status, "error")

        # 2) 纯文本全局重复
        for col_idx in range(len(headers)):
            if (row_idx, col_idx) in global_text_dups:
                h = headers[col_idx]
                val = global_text_dups[(row_idx, col_idx)]
                issues.append({
                    "row": excel_row_num,
                    "field": h,
                    "issue_type": "text_global_duplicate",
                    "original_value": val,
                    "suggestion": "该文本值（≥8字符）在表格中多处出现，请确认是否为复制粘贴错误",
                })
                row_status = _worse_status(row_status, "warning")

        # 3) 整行重复
        if row_idx in dup_row_indices:
            issues.append({
                "row": excel_row_num,
                "field": "(整行)",
                "issue_type": "row_duplicate",
                "original_value": "",
                "suggestion": "该行与表格中另一行完全相同（规范化后），将被跳过",
            })
            row_status = "duplicate"

        result.row_results.append({
            "row_index": excel_row_num,
            "row_status": row_status,
            "row_fingerprint": row_fingerprints[row_idx],
            "issues": issues,
            "raw_data": {h: _serialize_cell(v) for h, v in zip(headers, row_vals)},
        })

        if row_status == "ok":
            result.ok_rows += 1
        elif row_status == "error":
            result.error_rows += 1
        elif row_status == "warning":
            result.warning_rows += 1
        elif row_status == "duplicate":
            result.dup_rows += 1

    return result


def _worse_status(current: str, new: str) -> str:
    """状态优先级：error > duplicate > warning > ok"""
    order = {"ok": 0, "warning": 1, "duplicate": 2, "error": 3}
    return current if order.get(current, 0) >= order.get(new, 0) else new


def _to_field_key(label: str) -> str:
    """将列头文字转换为 snake_case field_key（用于字段注册表对比）。"""
    s = unicodedata.normalize("NFC", label.strip())
    # 将非字母数字替换为 _
    s = re.sub(r"[^\w]", "_", s, flags=re.UNICODE)
    s = re.sub(r"_+", "_", s).strip("_").lower()
    return s or "unknown"


def _serialize_cell(v: Any) -> Any:
    """将单元格值序列化为 JSON 兼容类型。"""
    if isinstance(v, datetime):
        return v.isoformat()
    if isinstance(v, date):
        return v.isoformat()
    if isinstance(v, time):
        return str(v)
    return v


# ===========================================================================
# 带颜色标注的 Excel 生成
# ===========================================================================

def generate_annotated_excel(
    file_bytes: bytes,
    preview_result: PreviewResult,
) -> bytes:
    """
    在原始 Excel 基础上叠加颜色标注，并附加 Issues sheet。
    返回标注后的 Excel bytes。
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes))
    ws = wb.active

    headers = preview_result.headers

    # 建立 (excel_row, col_idx) → [issue_types] 索引，方便快速查找
    cell_issues: dict[tuple[int, int], list[str]] = {}
    row_has_dup: set[int] = set()

    for row_res in preview_result.row_results:
        excel_row = row_res["row_index"]
        for issue in row_res.get("issues", []):
            itype = issue["issue_type"]
            field = issue["field"]
            if field == "(整行)":
                row_has_dup.add(excel_row)
                continue
            if field in headers:
                col_idx = headers.index(field)
                key = (excel_row, col_idx)
                cell_issues.setdefault(key, []).append(itype)

    # 给工作表单元格上色
    for row_res in preview_result.row_results:
        excel_row = row_res["row_index"]
        is_dup_row = excel_row in row_has_dup

        for col_idx in range(len(headers)):
            cell = ws.cell(row=excel_row, column=col_idx + 1)
            key = (excel_row, col_idx)
            ctypes = cell_issues.get(key, [])

            if "type_conflict" in ctypes:
                cell.fill = _FILL_RED
            elif "text_global_duplicate" in ctypes:
                cell.fill = _FILL_ORANGE
            elif is_dup_row:
                cell.fill = _FILL_YELLOW

    # 整行重复时，把整行标黄（已在上面逐列处理，但未覆盖没有 type/text 问题的列）
    for excel_row in row_has_dup:
        for col_idx in range(len(headers)):
            cell = ws.cell(row=excel_row, column=col_idx + 1)
            # 只有当前没有更高优先级颜色才上黄
            key = (excel_row, col_idx)
            ctypes = cell_issues.get(key, [])
            if not ctypes:
                cell.fill = _FILL_YELLOW

    # ── 生成 Issues sheet ────────────────────────────────────────────────────
    if "Issues" in wb.sheetnames:
        del wb["Issues"]
    ws_issues = wb.create_sheet("Issues")

    issue_headers = ["行号/列号", "字段名", "问题类型", "原值", "说明/建议处理方式"]
    ws_issues.append(issue_headers)
    # 表头加粗背景
    header_fill = PatternFill("solid", fgColor="D9D9D9")
    for col_idx, _ in enumerate(issue_headers, 1):
        c = ws_issues.cell(row=1, column=col_idx)
        c.fill = header_fill

    issue_type_cn = {
        "type_conflict": "类型冲突",
        "text_global_duplicate": "纯文本全局重复",
        "row_duplicate": "整行重复",
        "duplicate_header": "重复列名",
        "file_parse_error": "文件解析错误",
        "no_header": "缺少表头",
        "row_limit_truncated": "超行数截断",
        "no_data": "无数据行",
    }
    issue_fill_map = {
        "type_conflict": _FILL_RED,
        "text_global_duplicate": _FILL_ORANGE,
        "row_duplicate": _FILL_YELLOW,
        "duplicate_header": _FILL_PURPLE,
    }

    # 先给重复列头所在的整列（表头行）上紫色
    dup_header_cols: set[int] = set()   # 1-based col numbers
    for item in preview_result.errors:
        if item.get("issue_type") == "duplicate_header":
            col_1based = item.get("col_index")
            if col_1based:
                dup_header_cols.add(col_1based)
    for col_1based in dup_header_cols:
        ws.cell(row=1, column=col_1based).fill = _FILL_PURPLE

    for row_res in preview_result.row_results:
        for issue in row_res.get("issues", []):
            itype = issue["issue_type"]
            row_data = [
                issue.get("row", ""),
                issue.get("field", ""),
                issue_type_cn.get(itype, itype),
                issue.get("original_value", ""),
                issue.get("suggestion", ""),
            ]
            ws_issues.append(row_data)
            issue_row_num = ws_issues.max_row
            fill = issue_fill_map.get(itype)
            if fill:
                ws_issues.cell(row=issue_row_num, column=3).fill = fill

    # 全局错误/警告（无行号的）— duplicate_header 已有 field/col_index，单独处理
    for item in preview_result.errors + preview_result.warnings:
        itype = item.get("issue_type", "")
        field_label = item.get("field", "")
        row_label = item.get("col_index", "")   # duplicate_header 用列号代替行号
        ws_issues.append([
            row_label,
            field_label,
            issue_type_cn.get(itype, itype),
            "",
            item.get("suggestion", ""),
        ])
        issue_row_num = ws_issues.max_row
        fill = issue_fill_map.get(itype)
        if fill:
            ws_issues.cell(issue_row_num, column=3).fill = fill

    # 列宽自适应（简单估算）
    for col in ws_issues.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=0)
        ws_issues.column_dimensions[get_column_letter(col[0].column)].width = min(
            max(max_len + 2, 10), 60
        )

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
