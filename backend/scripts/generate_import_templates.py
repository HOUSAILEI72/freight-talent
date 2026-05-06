"""
生成 Excel 导入模板：岗位 / 候选人

用法：
  cd backend
  python scripts/generate_import_templates.py

输出：
  backend/templates/job_template.xlsx
  backend/templates/candidate_template.xlsx
"""
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT_DIR = Path(__file__).resolve().parent.parent / "templates"
OUT_DIR.mkdir(parents=True, exist_ok=True)

HEADER_FONT_FIXED = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
HEADER_FONT_TAG   = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
FILL_FIXED        = PatternFill("solid", fgColor="1F4E78")   # 蓝：固定字段
FILL_TAG          = PatternFill("solid", fgColor="2E7D32")   # 绿：自由标签列
HEADER_ALIGN      = Alignment(horizontal="center", vertical="center", wrap_text=True)
DATA_ALIGN        = Alignment(horizontal="left",   vertical="top",    wrap_text=True)
THIN              = Side(style="thin", color="DDDDDD")
BORDER            = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_header(cell, is_fixed=True):
    cell.font = HEADER_FONT_FIXED if is_fixed else HEADER_FONT_TAG
    cell.fill = FILL_FIXED if is_fixed else FILL_TAG
    cell.alignment = HEADER_ALIGN
    cell.border = BORDER


def style_data(cell):
    cell.alignment = DATA_ALIGN
    cell.border = BORDER


def write_notes_sheet(ws, notes_rows):
    for r, row in enumerate(notes_rows, start=1):
        for c, val in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=val)
            if r == 1:
                style_header(cell)
            else:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = BORDER


# ── 岗位模板 ──────────────────────────────────────────────────────────────────

def make_job_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "岗位"

    fixed = ["岗位名称", "省市区", "工作经验", "学历", "薪资", "职位详情", "岗位要求"]
    sample_tag_cols = ["板块", "业务类型", "语言要求"]
    headers = fixed + sample_tag_cols

    for col, h in enumerate(headers, start=1):
        style_header(ws.cell(row=1, column=col, value=h), is_fixed=(h in fixed))

    samples = [
        [
            "货代专员",
            "上海市-上海市-浦东新区",
            "3年以上",
            "本科",
            "8000-15000",
            "1. 处理客户订舱、报关、清关业务\n2. 跟踪货物动态并及时反馈\n3. 维护客户关系",
            "1. 大专以上学历\n2. 熟悉报关、报检流程\n3. 英语 CET-4 以上",
            "海运", "整柜", "英语",
        ],
        [
            "操作主管",
            "广东省-深圳市-福田区",
            "5年以上",
            "本科",
            "12000-20000",
            "1. 管理操作团队 6-10 人\n2. 把控订单时效与异常处理\n3. 优化操作流程",
            "1. 5年以上货代经验\n2. 团队管理经验 2 年以上\n3. 熟悉海空运操作",
            "空运", "拼箱", "英语",
        ],
        [
            "海外销售",
            "浙江省-宁波市-江北区",
            "1年以上",
            "本科",
            "面议+提成",
            "拓展海外货代客户，对接 NVOCC 与船公司",
            "1. 英语口语流利\n2. 抗压能力强\n3. 有海外驻外经验优先",
            "海运,空运", "整柜,拼箱", "英语,西班牙语",
        ],
    ]
    for r, row in enumerate(samples, start=2):
        for c, val in enumerate(row, start=1):
            style_data(ws.cell(row=r, column=c, value=val))

    widths = [12, 26, 10, 8, 14, 40, 40, 14, 14, 16]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 28
    for r in range(2, 5):
        ws.row_dimensions[r].height = 80

    ws.freeze_panes = "A2"

    notes = [
        ["字段", "说明", "示例"],
        ["岗位名称", "必填，岗位的显示名称", "货代专员"],
        ["省市区", "必填。用 - 分隔三级行政区，系统会拆开供分别筛选", "上海市-上海市-浦东新区"],
        ["工作经验", "建议值: 不限 / 应届 / 1年以上 / 3年以上 / 5年以上 / 10年以上", "3年以上"],
        ["学历", "建议值: 不限 / 高中 / 大专 / 本科 / 硕士 / 博士", "本科"],
        ["薪资", "自由文本，可填范围或描述", "8000-15000 或 面议+提成"],
        ["职位详情", "岗位职责。多行用换行（Alt+Enter）", "1. ...\n2. ..."],
        ["岗位要求", "任职要求。多行用换行（Alt+Enter）", "1. ...\n2. ..."],
        ["", "", ""],
        ["📌 自由标签列说明", "", ""],
        ["其他任意列", "列名将作为「分类」，单元格值将作为「标签」。同一格里可写多个标签，用 , 或 、 分隔。", "板块=海运,空运"],
        ["新分类首次出现", "进入审批中心待管理员通过后，才会出现在筛选器", ""],
        ["列名相同视为同一分类", "如果跨次上传都用 \"语言要求\" 这个列名，会合并到同一分类下", ""],
    ]
    ws2 = wb.create_sheet("说明")
    write_notes_sheet(ws2, notes)
    ws2.column_dimensions["A"].width = 24
    ws2.column_dimensions["B"].width = 60
    ws2.column_dimensions["C"].width = 30

    out = OUT_DIR / "job_template.xlsx"
    wb.save(out)
    return out


# ── 候选人模板 ────────────────────────────────────────────────────────────────

def make_candidate_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "候选人"

    fixed = [
        "姓名", "年龄", "工作年限", "学历", "电话号码", "求职状态",
        "工作经历", "教育经历", "资格证书",
    ]
    sample_tag_cols = ["意向城市", "板块", "业务类型", "语言能力"]
    headers = fixed + sample_tag_cols

    for col, h in enumerate(headers, start=1):
        style_header(ws.cell(row=1, column=col, value=h), is_fixed=(h in fixed))

    samples = [
        [
            "张三", 28, 5, "本科", "13800001234", "open",
            "2020-2024 | 高级货代员 | ABC物流有限公司\n2018-2020 | 货代操作 | XYZ国际货运",
            "2014-2018 | 上海海事大学 | 物流管理 | 本科",
            "国际货运代理资格证、报关员资格证",
            "上海", "海运", "整柜", "英语",
        ],
        [
            "李四", 32, 8, "硕士", "13900005678", "passive",
            "2019-至今 | 操作主管 | DEF船代\n2015-2019 | 资深操作 | GHI物流\n2012-2015 | 操作专员 | JKL货运",
            "2013-2015 | 中山大学 | 物流工程 | 硕士\n2009-2013 | 暨南大学 | 国际经济与贸易 | 本科",
            "报关员、海事英语证书",
            "深圳,广州", "空运", "拼箱", "英语,日语",
        ],
        [
            "王五", 26, 3, "本科", "13700009999", "open",
            "2021-至今 | 海外销售 | MNO国际",
            "2017-2021 | 大连海事大学 | 航海技术 | 本科",
            "",
            "宁波", "海运,空运", "整柜", "英语",
        ],
    ]
    for r, row in enumerate(samples, start=2):
        for c, val in enumerate(row, start=1):
            style_data(ws.cell(row=r, column=c, value=val))

    widths = [10, 6, 8, 8, 14, 10, 50, 50, 28, 14, 14, 14, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 28
    for r in range(2, 5):
        ws.row_dimensions[r].height = 90

    ws.freeze_panes = "A2"

    notes = [
        ["字段", "说明", "示例"],
        ["姓名", "必填", "张三"],
        ["年龄", "整数", "28"],
        ["工作年限", "整数，单位「年」", "5"],
        ["学历", "建议值: 高中 / 大专 / 本科 / 硕士 / 博士", "本科"],
        ["电话号码", "11 位手机号", "13800001234"],
        ["求职状态", "open=开放机会 / passive=被动寻找 / closed=暂不考虑", "open"],
        ["工作经历", "每行一条经历，字段用 | 分隔：\n年限 | 职位 | 公司\n多条之间用换行（Alt+Enter）",
         "2020-2024 | 高级货代员 | ABC物流"],
        ["教育经历", "每行一条经历，字段用 | 分隔：\n年限 | 院校 | 专业 | 学历（学历可省略）\n多条之间用换行",
         "2014-2018 | 上海海事大学 | 物流管理 | 本科"],
        ["资格证书", "多个证书用顿号、逗号或换行分隔", "国际货代证、报关员"],
        ["", "", ""],
        ["📌 自由标签列说明", "", ""],
        ["其他任意列", "列名作为「分类」，单元格值作为「标签」，多个值用 , 或 、 分隔",
         "板块=海运,空运"],
        ["", "", ""],
        ["⚠️ 隐私规则", "", ""],
        ["以下字段在企业方默认隐藏：", "姓名 / 年龄 / 工作年限 / 学历 / 电话号码 / 求职状态 / 工作经历 / 教育经历 / 资格证书", ""],
        ["何时对企业可见？", "候选人接受了该企业的邀约后，仅该企业可见这些字段；其他企业仍然只看到匿名标签", ""],
        ["公开可见的内容", "候选人详情页与列表只展示「自由标签列」的标签（如：意向城市、板块、语言能力等），不暴露个人信息",
         ""],
    ]
    ws2 = wb.create_sheet("说明")
    write_notes_sheet(ws2, notes)
    ws2.column_dimensions["A"].width = 26
    ws2.column_dimensions["B"].width = 64
    ws2.column_dimensions["C"].width = 30

    out = OUT_DIR / "candidate_template.xlsx"
    wb.save(out)
    return out


if __name__ == "__main__":
    p1 = make_job_template()
    p2 = make_candidate_template()
    print(f"[OK] Job template     -> {p1}")
    print(f"[OK] Resume template  -> {p2}")
